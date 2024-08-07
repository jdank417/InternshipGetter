import subprocess
import requests
import spacy
import logging
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Excel file setup
EXCEL_FILE = 'internshipGetter.xlsx'
SHEET_NAME = 'Sheet1'

def initialize_excel_sheet(excel_file, sheet_name):
    try:
        wb = load_workbook(excel_file)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['Title', 'Link'])  # Add headers
            wb.save(excel_file)
            logging.info(f"Created new sheet with headers.")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(['Title', 'Link'])  # Add headers
        wb.save(excel_file)
        logging.info(f"Created new Excel file with headers.")

def get_existing_jobs(excel_file, sheet_name):
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    if 'Title' not in df.columns:
        raise KeyError("The 'Title' column is missing in the Excel sheet.")
    return df['Title'].tolist()

def add_job_to_excel(excel_file, sheet_name, job_title, job_link):
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    new_row = pd.DataFrame([[job_title, job_link]], columns=['Title', 'Link'])
    df = pd.concat([df, new_row], ignore_index=True)
    df.to_excel(excel_file, sheet_name=sheet_name, index=False)
    logging.info(f"Added job to Excel: {job_title}")
    autosize_columns(excel_file, sheet_name)

def autosize_columns(excel_file, sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(excel_file)

# spaCy model setup
try:
    nlp = spacy.load('en_core_web_sm')
except OSError:
    subprocess.run(['python', '-m', 'spacy', 'download', 'en_core_web_sm'])
    nlp = spacy.load('en_core_web_sm')

# Read resume content from file
try:
    with open('resume.txt', 'r') as file:
        resume_text = file.read()
except FileNotFoundError:
    logging.error("Error: 'resume.txt' not found.")
    resume_text = ""

def extract_keywords(text):
    doc = nlp(text.lower())  # Lowercase text to improve matching accuracy
    return [token.lemma_ for token in doc if token.is_alpha and not token.is_stop]

def match_jobs_with_resume(job_descriptions, resume_text):
    if not resume_text:
        return []

    resume_keywords = extract_keywords(resume_text)
    job_matches = []

    for job in job_descriptions:
        job_text = job['title'] + " " + job['description']
        job_keywords = extract_keywords(job_text)
        combined_text = [" ".join(resume_keywords), " ".join(job_keywords)]

        vectorizer = TfidfVectorizer().fit_transform(combined_text)
        similarity = cosine_similarity(vectorizer[0:1], vectorizer[1:2]).flatten()[0]

        job_matches.append((similarity, job))

    # Sort by similarity score
    job_matches.sort(reverse=True, key=lambda x: x[0])
    return job_matches


def search_internships(api_key, cse_id, query, location, resume_text, excel_file, sheet_name, num_results=10):
    query_with_location = f"{query} in {location}"
    start_index = 1
    all_job_descriptions = []

    while len(all_job_descriptions) < num_results:
        url = f"https://www.googleapis.com/customsearch/v1?q={query_with_location}&key={api_key}&cx={cse_id}&start={start_index}"
        response = requests.get(url)

        if response.status_code == 200:
            results = response.json().get('items', [])
            if not results:
                logging.info("No more job descriptions found.")
                break

            job_descriptions = [{'title': item['title'], 'link': item['link'], 'description': item['snippet']} for item
                                in results]
            all_job_descriptions.extend(job_descriptions)

            start_index += 10  # Move to the next page
        else:
            logging.error(f"Error: {response.status_code} - {response.text}")
            break

    # Limit the number of job descriptions to the specified number of results
    all_job_descriptions = all_job_descriptions[:num_results]

    if not all_job_descriptions:
        logging.info("No job descriptions found.")
        return

    existing_jobs = get_existing_jobs(excel_file, sheet_name)
    matched_jobs = match_jobs_with_resume(all_job_descriptions, resume_text)

    for score, job in matched_jobs:
        if job['title'] not in existing_jobs:
            add_job_to_excel(excel_file, sheet_name, job['title'], job['link'])
            print(
                f"Relevance: {score:.2f}\nTitle: {job['title']}\nLink: {job['link']}\nDescription: {job['description']}\n")
        else:
            logging.info(f"Job already listed: {job['title']}")


# Configuration
api_key = "AIzaSyCmDTs2lWUqNrQqRCEorKI1umwTdRpcAqM"
cse_id = "61fcd2bdd7a354817"
query = "software engineering internship Spring 2025"
location = "Boston, MA"
excel_file = EXCEL_FILE
sheet_name = SHEET_NAME

# Initialize Excel file and sheet if necessary
initialize_excel_sheet(excel_file, sheet_name)

search_internships(api_key, cse_id, query, location, resume_text, excel_file, sheet_name, num_results=25)
